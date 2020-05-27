import random
import time
from decimal import *
import numpy
from datetime import datetime
from openpyxl import load_workbook
from DMN import *
import matplotlib.pyplot as plt


def calc_pos(ind, col_count):
    ind = ind
    row = ind / col_count
    col = ind % col_count
    return {'row': row, 'col': col}


def main():
    index = 0
    array = random.randint(0, 100)
    print(array)
    wb3 = load_workbook(filename="mirna_drug.xlsx", read_only=True)
    ws3 = wb3["Sheet1"]
    y_row_count = ws3.max_row - 1
    y_col_count = ws3.max_column - 1

    y_matrix = numpy.zeros((y_row_count, y_col_count))
    y_matrix_t = numpy.transpose(y_matrix)
    x_count = 0
    i = 0
    for row in ws3.rows:
        if (i > 0):  # skip first row
            for j in range(y_col_count):
                j += 1  # skip first column
                if (row[j].value):
                    y_matrix[i - 1][j - 1] = row[j].value  # (i-1) and (j-1) indexes used for skipped first row and col
        i += 1

    wb = load_workbook(filename="mirna_sim.xlsx", read_only=True)
    ws = wb["Sheet1"]
    sm_row_count = ws.max_row - 1

    sm_matrix = numpy.zeros((sm_row_count, sm_row_count))
    dm_matrix = numpy.zeros((sm_row_count, sm_row_count))
    x_count = 0
    i = 0
    for row in ws.rows:
        if (i > 0):  # skip first row
            for j in range(sm_row_count):
                j += 1  # skip first column
                if (row[j].value):
                    sm_matrix[i - 1][j - 1] = row[j].value  # (i-1) and (j-1) indexes used for skipped first row and col
        i += 1

    for x in range(sm_row_count):
        sm_x = sm_matrix[x]
        for y in range(sm_row_count):
            x_count += sm_x[y]
        dm_matrix[x][x] = Decimal(x_count)
        x_count = 0

    wb2 = load_workbook(filename="drug_sim.xlsx", read_only=True)
    ws2 = wb2["Sheet1"]
    su_row_count = ws2.max_row - 1

    su_matrix = numpy.zeros((su_row_count, su_row_count))
    du_matrix = numpy.zeros((su_row_count, su_row_count))
    x_count = 0
    i = 0
    for row in ws2.rows:
        if (i > 0):  # skip first row
            for j in range(su_row_count):
                j += 1  # skip first column
                if (row[j].value):
                    su_matrix[i - 1][j - 1] = row[j].value  # (i-1) and (j-1) indexes used for skipped first row and col
        i += 1
    for x in range(su_row_count):
        su_x = su_matrix[x]
        for y in range(su_row_count):
            x_count += su_x[y]
        du_matrix[x][x] = Decimal(x_count)
        x_count = 0

    eta_list = [0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8, 0.9, 1]
    thr_list = [0.01, 0.02, 0.03, 0.04, 0.05, 0.06, 0.07, 0.08, 0.09, 0.1, 0.11, 0.12, 0.13, 0.14, 0.15, 0.16, 0.17,
                0.18, 0.19, 0.2, 0.21, 0.22, 0.23, 0.24, 0.25, 0.26, 0.27, 0.28, 0.29, 0.3, 0.31, 0.32, 0.33, 0.34,
                0.35, 0.36, 0.37, 0.38, 0.39, 0.4, 0.41, 0.42, 0.43, 0.44, 0.45, 0.46, 0.47, 0.48, 0.49, 0.5, 0.51,
                0.52, 0.53, 0.54, 0.55, 0.56, 0.57, 0.58, 0.59, 0.6, 0.61, 0.62, 0.63, 0.64, 0.65, 0.66, 0.67, 0.68,
                0.69, 0.7, 0.71, 0.72, 0.73, 0.74, 0.75, 0.76, 0.77, 0.78, 0.79, 0.8, 0.81, 0.82, 0.83, 0.84, 0.85,
                0.86, 0.87, 0.88, 0.89, 0.9, 0.91, 0.92, 0.93, 0.94, 0.95, 0.96, 0.97, 0.98, 0.99]
    K_list = [1,2,3,4,5,6,7,8,9,10]
    k_list = [50,100,150]
    lambda_l_list = [0.25,0.5,1,2] 
    lambda_m_list = [0,0.0001,0.001,0.01,0.1]
    cv_average = []
    #### begin of cross validation folding
    fold_count = 10

    one_element_array = DMN_N()
    one_element_array = list(map(int, one_element_array))
    # one_element_array = list(range(0, y_row_count * y_col_count))
    random.shuffle(one_element_array)
    random_indexes = numpy.array(one_element_array)

    folds_index = numpy.array_split(random_indexes, fold_count)

    cross_validation_result_file = open(
        "cross_validation_by_element_fix_thr_result_" + time.strftime("%Y%m%d-%H%M%S") + ".csv", "w")
    cross_validation_result_file.write(
        'fold_number,K,eta,k,lambda_l,lambda_m,thr,TP,TN,FP,FN,MCC,ACC,Sensitivity(TPR),Specificity(TNR),Precision(PPV),NPV,FNR,FPR,FDR,FOR,TS,F1,Informedness,Markedness\n')
    TP = float(1)
    FN = float(1)
    TN = float(1)
    FP = float(1)
    cross_validation_average_file = open(
        "cross_validation_by_element_Average_" + time.strftime("%Y%m%d-%H%M%S") + ".csv", "w")
    cross_validation_average_file.write(
        'K,eta,k,lambda_l,lambda_m,thr,MCC,ACC,Sensitivity(TPR),Specificity(TNR),Precision(PPV),NPV,FNR,FPR,FDR,FOR,TS,F1,Informedness,Markedness,auc,auc_sum\n')
    tp_sum = 0
    tn_sum = 0
    fp_sum = 0
    fn_sum = 0
    thr_value = numpy.zeros((14, len(thr_list)))
    # Avg_TP = 0
    # Avg_TN = 0
    # Avg_FP = 0
    # Avg_FN = 0
    for f in range(fold_count):
        y_matrix_folded = numpy.copy(y_matrix)
        y_matrix_folded_none_zero = numpy.copy(y_matrix)
        for ind in folds_index[f]:
            result = calc_pos(ind, y_col_count)
            y_matrix_folded[int(result['row']), result['col']] = 0

        for K in K_list:
            for eta1 in eta_list:
                # Using neighbor information of drugs
                eta = eta1 ** (numpy.arange(0, K))
                y2_new1 = numpy.zeros_like(y_matrix_folded)
                empty_rows = numpy.where(~y_matrix_folded.any(axis=1))[0]
                empty_cols = numpy.where(~y_matrix_folded.any(axis=0))[0]
                for i in numpy.arange(0, max(numpy.shape(sm_matrix))):
                    drug_sim = sm_matrix[i, :]
                    drug_sim[i] = 0
                    indices = numpy.arange(0, max(numpy.shape(sm_matrix)))  # row or col
                    drug_sim = numpy.delete(drug_sim, empty_rows)
                    indices = numpy.delete(indices, empty_rows)
                    indx = numpy.argsort(-drug_sim)
                    indx = indx[0:K]
                    indx = indices[indx]
                    drug_sim = sm_matrix[i, :]

                    aa = eta * drug_sim[indx]
                    yy = y_matrix_folded[indx, :]
                    bb = numpy.dot(aa, yy)
                    cc = sum(drug_sim[indx])
                    if (cc != 0):
                        dd = bb / cc
                    else:
                        dd = 0
                    y2_new1[i, :] = dd
				
				# Using neighbor information of microRNAs
                y2_new2 = numpy.zeros_like(y_matrix_folded)
                for j in numpy.arange(0, max(numpy.shape(su_matrix))):
                    target_sim = numpy.array(su_matrix[j, :])
                    target_sim[j] = 0
                    indices = numpy.arange(0, max(numpy.shape(su_matrix)))
                    target_sim = numpy.delete(target_sim, empty_cols)
                    indices = numpy.delete(indices, empty_cols)
                    indx = numpy.argsort(-target_sim)
                    indx = indx[0:K]
                    indx = indices[indx]
                    target_sim = su_matrix[j, :]

                    aa = numpy.transpose(eta * target_sim[indx])
                    yy = y_matrix_folded[:, indx]
                    bb = numpy.dot(yy, aa)
                    cc = sum(target_sim[indx])
                    if (cc != 0):
                        dd = bb / cc
                    else:
                        dd = 0
                    y2_new2[:, j] = dd

                y2_avg = (y2_new1 + y2_new2) / 2
                Y = numpy.maximum(y_matrix_folded, y2_avg)
                Y_t = numpy.transpose(Y)


                for k in k_list:
                    for lambda_l in lambda_l_list:
                        for lambda_m in lambda_m_list:
                            W = numpy.random.uniform(low=0, high=1, size=(y_row_count, k))
                            H = numpy.random.uniform(low=0, high=1, size=(y_col_count, k))
                            counterw = 0
                            counterh = 0
                            w_cell_count = 0
                            h_cell_count = 0
                            flag = True
                            iteratew = 0
                            iterateh = 0
                            lambda_m_sm = lambda_m * sm_matrix
                            lambda_m_su = lambda_m * su_matrix
                            lambda_m_dm = lambda_m * dm_matrix
                            lambda_m_du = lambda_m * du_matrix
                            while flag == True:
                                W_old = numpy.copy(W)
                                wn1 = numpy.matmul(Y, H)
                                w_numerator = wn1 + numpy.matmul(lambda_m_sm, W)
                                wd11 = numpy.matmul(W, numpy.transpose(H))
                                w_denominator = (numpy.matmul(wd11, H)) + (lambda_l * W) + (
                                    numpy.matmul(lambda_m_dm, W))
                                # H
                                H_old = numpy.copy(H)
                                hn1 = numpy.matmul(Y_t, W)
                                hn2 = numpy.matmul(lambda_m_su, H)
                                h_numerator = hn1 + hn2
                                hd11 = numpy.matmul(H, numpy.transpose(W))
                                h_denominator = (numpy.matmul(hd11, W)) + (lambda_l * H) + (
                                    numpy.matmul(lambda_m_du, H))

                                for i in range(y_row_count):
                                    for j in range(k):
                                        wik = W[i][j]
                                        wik = wik * (w_numerator[i][j] / w_denominator[i][j])
                                        W[i][j] = wik
                                w_difference = numpy.absolute(W - W_old)

                                for i in range(y_col_count):
                                    for j in range(k):
                                        hjk = H[i][j]
                                        hjk = hjk * (h_numerator[i][j] / h_denominator[i][j])
                                        H[i][j] = hjk
                                h_difference = numpy.absolute(H - H_old)
                                for i in range(y_row_count):
                                    for j in range(k):
                                        counterw += 1
                                        if w_difference[i][j] <= 0.0001:
                                            w_cell_count += 1
                                            if w_cell_count == (y_row_count * k):
                                                iteratew = 1
                                                # print 'count-w=', counterw
                                            elif i == y_row_count:
                                                w_cell_count = 0
                                for i in range(y_col_count):
                                    for j in range(k):
                                        counterh += 1
                                        if h_difference[i][j] <= 0.0001:
                                            h_cell_count += 1
                                            if h_cell_count == (y_col_count * k):
                                                iterateh = 1
                                                # print 'count-h=', counterh
                                            elif i == y_col_count:
                                                h_cell_count = 0
                                if iteratew == 1 and iterateh == 1:
                                    if iterateh == 1:
                                        flag = False
                            print("k = ", k, ",lambda_l = ", lambda_l, ",lambda_m = ", lambda_m,
                                  "    This is printed on: ", str(datetime.now()))
                            YStar = numpy.matmul(W, numpy.transpose(H))
                            YStar_array = numpy.array(YStar).flatten()
                            numpy.savetxt(
                                'YStar(K=' + str(K) + ',eta=' + str(eta1) + ',k=' + str(k) + ',lambda_l=' + str(
                                    lambda_l) + ',lambda_m=' + str(lambda_m) + ').txt', YStar, '%0.5f')

                            numpy.savetxt(
                                'YStar_array(K=' + str(K) + ',eta=' + str(eta1) + ',k=' + str(
                                    k) + ',lambda_l=' + str(
                                    lambda_l) + ',lambda_m=' + str(lambda_m) + ').txt', YStar_array, '%0.5f')

                            count = 0
                            for thr in thr_list:
                                TP = float(1)
                                TN = float(1)
                                FN = float(1)
                                FP = float(1)
                                thr_none_zero = []
                                thr_zero = []
                                YStar_result = []
                                test_result_zero = []
                                test_result_none_zero = []
                                for ind in folds_index[f]:
                                    result = calc_pos(ind, y_col_count)
                                    YStar_result.append(YStar[int(result['row']), result['col']])
                                    test_result_zero.append(y_matrix_folded[int(result['row']), result['col']])
                                    test_result_none_zero.append(
                                        y_matrix_folded_none_zero[int(result['row']), result['col']])

                                positive_index = []
                                negative_index = []
                                index = 0
                                for tnz in test_result_none_zero:
                                    if (tnz == 1):
                                        positive_index.append(index)
                                    else:
                                        negative_index.append(index)
                                    index = index + 1

                                negative_index = random.sample(set(negative_index), len(positive_index))

                                for tz_ind, tnz_ind in zip(negative_index, positive_index):
                                    if (test_result_none_zero[tnz_ind] == 1):
                                        if (YStar_result[tnz_ind] >= thr):
                                            TP += 1

                                        else:
                                            FN += 1

                                    else:
                                        if (YStar_result[tnz_ind] >= thr):
                                            FP += 1

                                        else:
                                            TN += 1

                                    if (test_result_none_zero[tz_ind] == 1):
                                        if (YStar_result[tz_ind] >= thr):
                                            TP += 1

                                        else:
                                            FN += 1

                                    else:
                                        if (YStar_result[tz_ind] >= thr):
                                            FP += 1

                                        else:
                                            TN += 1

                                        MCC = ((TP * TN) - (FP * FN)) / (
                                            numpy.sqrt((TP + FP) * (TP + FN) * (TN + FP) * (TN + FN)))
                                        ACC = (TP + TN) / (TP + TN + FP + FN)
                                        TPR = TP / (TP + FN)
                                        TNR = TN / (TN + FP)
                                        PPV = TP / (TP + FP)
                                        NPV = TN / (TN + FN)
                                        FNR = FN / (FN + TP)
                                        FPR = FP / (FP + TN)
                                        FDR = FP / (FP + TP)
                                        FOR = FN / (FN + TN)
                                        TS = TP / (TP + FN + FP)
                                        F1 = (2 * TP) / (2 * TP + FP + FN)
                                        BM = TPR + TNR - 1
                                        MK = PPV + NPV - 1

                                cv_average.append([str(K) + ',' + str(eta1) + ',' + str(k) + ',' + str(
                                    lambda_l) + ',' + str(lambda_m) + ',' + str(thr), str(f), str(K), str(eta1), str(k),
                                                   str(lambda_l), str(lambda_m), str(thr), str(MCC), str(ACC), str(TPR),
                                                   str(TNR), str(PPV), str(NPV), str(FNR), str(FPR), str(FDR), str(FOR),
                                                   str(TS), str(F1), str(BM), str(MK), str(0), str(0)])

                                cross_validation_result_file.write(
                                    str(f) + "," +
                                    str(K) + "," +
                                    str(eta1) + "," +
                                    str(k) + "," +
                                    str(lambda_l) + "," +
                                    str(lambda_m) + "," +
                                    str(thr) + "," +
                                    str(TP) + "," +
                                    str(TN) + "," +
                                    str(FP) + "," +
                                    str(FN) + "," +
                                    str(MCC) + "," +
                                    str(ACC) + "," +
                                    str(TPR) + "," +
                                    str(TNR) + "," +
                                    str(PPV) + "," +
                                    str(NPV) + "," +
                                    str(FNR) + "," +
                                    str(FPR) + "," +
                                    str(FDR) + "," +
                                    str(FOR) + "," +
                                    str(TS) + "," +
                                    str(F1) + "," +
                                    str(BM) + "," +
                                    str(MK) + "\n")

    # end of fold

    cv_average = sorted(cv_average)
    fpr_tpr_array = []
    fpr_tpr_array.append([1, 1])
    auc_array = []
    i = 1
    auc_sum = 0
    for x in range(len(thr_list) * len(K_list) * len(eta_list) * len(k_list) * len(lambda_l_list) * len(lambda_m_list)):
        sum_MCC = 0
        sum_ACC = 0
        sum_TPR = 0
        sum_TNR = 0
        sum_PPV = 0
        sum_NPV = 0
        sum_FNR = 0
        sum_FPR = 0
        sum_FDR = 0
        sum_FOR = 0
        sum_TS = 0
        sum_F1 = 0
        sum_BM = 0
        sum_MK = 0
        for y in range((x * fold_count), (x * fold_count) + 10):
            sum_MCC += float(cv_average[y][8])
            sum_ACC += float(cv_average[y][9])
            sum_TPR += float(cv_average[y][10])
            sum_TNR += float(cv_average[y][11])
            sum_PPV += float(cv_average[y][12])
            sum_NPV += float(cv_average[y][13])
            sum_FNR += float(cv_average[y][14])
            sum_FPR += float(cv_average[y][15])
            sum_FDR += float(cv_average[y][16])
            sum_FOR += float(cv_average[y][17])
            sum_TS += float(cv_average[y][18])
            sum_F1 += float(cv_average[y][19])
            sum_BM += float(cv_average[y][20])
            sum_MK += float(cv_average[y][21])
        average_MCC = sum_MCC / fold_count
        average_ACC = sum_ACC / fold_count
        average_TPR = sum_TPR / fold_count
        average_TNR = sum_TNR / fold_count
        average_PPV = sum_PPV / fold_count
        average_NPV = sum_NPV / fold_count
        average_FNR = sum_FNR / fold_count
        average_FPR = sum_FPR / fold_count
        average_FDR = sum_FDR / fold_count
        average_FOR = sum_FOR / fold_count
        average_TS = sum_TS / fold_count
        average_F1 = sum_F1 / fold_count
        average_BM = sum_BM / fold_count
        average_MK = sum_MK / fold_count

        fpr_tpr_array.append([average_FPR, average_TPR])
        fpr_tpr_i = fpr_tpr_array[i - 1]
        fpr_tpr_i_2 = fpr_tpr_array[i]
        auc_value = 0.5 * (fpr_tpr_i[0] - fpr_tpr_i_2[0]) * (fpr_tpr_i[1] + fpr_tpr_i_2[1])
        auc_array.append(auc_value)
        auc_sum += auc_value
        re = x % len(thr_list)
        i += 1

        # for i in range(len(thr_list)):
        cross_validation_average_file.write(
            str(cv_average[x * fold_count][2]) + "," +
            str(cv_average[x * fold_count][3]) + "," +
            str(cv_average[x * fold_count][4]) + "," +
            str(cv_average[x * fold_count][5]) + "," +
            str(cv_average[x * fold_count][6]) + "," +
            str(thr_list[re]) + "," +
            str(average_MCC) + "," +
            str(average_ACC) + "," +
            str(average_TPR) + "," +
            str(average_TNR) + "," +
            str(average_PPV) + "," +
            str(average_NPV) + "," +
            str(average_FNR) + "," +
            str(average_FPR) + "," +
            str(average_FDR) + "," +
            str(average_FOR) + "," +
            str(average_TS) + "," +
            str(average_F1) + "," +
            str(average_BM) + "," +
            str(average_MK) + "," +
            str(auc_value) + "," +
            str(auc_sum) + "\n")
        # print x

    cross_validation_result_file.close()
    cross_validation_average_file.close()
    fpt_tpr_array_t = numpy.transpose(fpr_tpr_array)
    plt.plot(fpt_tpr_array_t[1], fpt_tpr_array_t[0])
    plt.ylabel('Sensivity')
    plt.xlabel('FPR')

    plt.savefig(
        fname='K=' + str(K) + ',eta=' + str(eta1) + ',k=' + str(k) + ',lambda_l=' + str(lambda_l) + ',lambda_m=' + str(
            lambda_m) + ').png')


if __name__ == '__main__':
    main()
